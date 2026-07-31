"""Fish Care integration tests — real Postgres (Docker) with the V0.1.6 migration applied.

Uses a synthetic multi-sheet xlsx built in-test (rather than the ~700KB real workbook) so row
counts are deterministic; it still exercises openpyxl parsing, replace-per-sheet persistence, the
read path, and the (mocked) Google Sheets adapter contract.
"""
from unittest.mock import patch

import pytest
from openpyxl import Workbook

from db_utils.core import ResultType, execute_statements
from importer import google_sheets
from importer.import_fish_care import parse_workbook, parse_xlsx_file
from model.fish_care import get_fish_care, persist_sheets
from species_config.lfs import CONFIG as LFS

USER = 'amphiadmin'

HEADER = ['Date', 'Tank ID', 'Carer', 'Temp', 'DO', 'Salinity', 'pH',
          'Ammonia', 'Nitrite', 'Nitrate', 'Morts', 'Notes']


@pytest.fixture
def clean_fish_care(set_cleanup_sql_fn):
    # Isolate each test, and leave the table clean for reused (fixed-container) runs.
    execute_statements("TRUNCATE fish_care", USER, ResultType.NoResult)
    set_cleanup_sql_fn("TRUNCATE fish_care")
    yield


def _add_sheet(wb, title, header, rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for row in rows:
        ws.append(row)


def _make_workbook(path, charlie_rows=2):
    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, 'Charlie 26', HEADER,
               [['2026-01-09', f'2{chr(65 + i)}', '', '11.2', '5.36', '7.6', '0.0', '', '', '', str(i), '']
                for i in range(charlie_rows)])
    _add_sheet(wb, 'Echo 26', HEADER,
               [['2026-01-24', 'C11', 'BY', '11.3', '10.4', '5.6', '', '', '', '', '0', 'Added larvae']])
    _add_sheet(wb, 'LFS Wet Lab 26 SYS 3', HEADER,
               [['2025-12-15', '3A', 'ZK', '11.3', '', '~5', '8.3', '', '', '', '1', 'fed'],
                ['2025-12-16', '3B', 'ZK', '11.4', '', '~5', '8.2', '', '', '', '2', '']])
    # Unrecognized maintenance sheet — must be skipped, not imported.
    _add_sheet(wb, 'Equipment 26', ['Date', 'Location', 'Item', 'Maintenance Performed'],
               [['2026-01-01', 'Charlie', 'pump', 'cleaned']])
    wb.save(path)


def test_import_counts_and_skips_equipment(tmp_path, clean_fish_care):
    path = str(tmp_path / 'fc.xlsx')
    _make_workbook(path)

    result = parse_workbook(parse_xlsx_file(path), LFS)
    assert result['skipped'] == ['Equipment 26']

    persist_sheets(result['parsed'], USER, 'xlsx')
    rows = get_fish_care(USER)
    assert len(rows) == 5  # Charlie 2 + Echo 1 + SYS3 2
    assert len([r for r in rows if r['facility'] == 'Charlie']) == 2
    # Approximate value preserved verbatim as text.
    assert any(r['salinity'] == '~5' for r in rows)


def test_reimport_replaces_no_duplicates(tmp_path, clean_fish_care):
    path = str(tmp_path / 'fc.xlsx')
    _make_workbook(path)

    persist_sheets(parse_workbook(parse_xlsx_file(path), LFS)['parsed'], USER, 'xlsx')
    assert len(get_fish_care(USER)) == 5

    # Same workbook again -> replace-per-sheet, still 5 rows.
    persist_sheets(parse_workbook(parse_xlsx_file(path), LFS)['parsed'], USER, 'xlsx')
    assert len(get_fish_care(USER)) == 5


def test_single_sheet_import_leaves_others_intact(tmp_path, clean_fish_care):
    full = str(tmp_path / 'full.xlsx')
    _make_workbook(full, charlie_rows=2)
    persist_sheets(parse_workbook(parse_xlsx_file(full), LFS)['parsed'], USER, 'xlsx')
    assert len(get_fish_care(USER)) == 5

    # Re-upload ONLY Charlie 26 with a single row.
    charlie_only = str(tmp_path / 'charlie.xlsx')
    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, 'Charlie 26', HEADER,
               [['2026-02-01', '2A', 'KA', '11.0', '5.0', '7.5', '0.0', '', '', '', '0', 'redo']])
    wb.save(charlie_only)
    persist_sheets(parse_workbook(parse_xlsx_file(charlie_only), LFS)['parsed'], USER, 'xlsx')

    rows = get_fish_care(USER)
    assert len([r for r in rows if r['facility'] == 'Charlie']) == 1  # replaced
    assert len([r for r in rows if r['facility'] == 'Echo']) == 1     # untouched
    assert len([r for r in rows if r['facility'] == 'LFS Wet Lab']) == 2  # untouched


def test_google_sheets_path_persists(clean_fish_care):
    # Mock the adapter's network call; assert the sync core parses + persists and tags source.
    canned = {
        'Charlie 26': (HEADER, [['2026-01-09', '2A', '', '11.2', '5.36', '7.6', '0.0', '', '', '', '0', '']]),
        'Equipment 26': (['Date', 'Location', 'Item'], [['2026-01-01', 'x', 'y']]),
    }
    with patch('importer.google_sheets.read_spreadsheet', return_value=canned):
        sheets = google_sheets.read_spreadsheet('fake-sheet-id')

    result = parse_workbook(sheets, LFS)
    out = persist_sheets(result['parsed'], USER, 'gsheet')
    assert out['success']['inserted']['fish_care'] == 1

    rows = get_fish_care(USER)
    assert len(rows) == 1
    assert rows[0]['source'] == 'gsheet'

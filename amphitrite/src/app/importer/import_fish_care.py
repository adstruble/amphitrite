"""Fish Care import — species-agnostic parse/persist core.

All species specifics (recognized facilities, header spellings) come from the SpeciesConfig; there
are no hardcoded facility or column names here. Two adapters feed the same core: an xlsx file
(openpyxl) and a Google Sheets pull (importer/google_sheets.py). Both build a `sheets` dict of
`{sheet_name: (header, data_rows)}` and call parse_workbook().
"""
import os
import re
from datetime import date, datetime

from amphi_logging.logger import get_logger
from model.fish_care import persist_sheets
from species_config import get_species_config
from utils.server_state import JobState, complete_job

LOGGER = get_logger('importer')

# Sheet name = facility prefix + 2-digit year + optional 'SYS n', e.g. 'LFS Wet Lab 26 SYS 3'.
_SHEET_SUFFIX_RE = re.compile(r'^(\d{2})(?:\s+SYS\s+(\S+))?$', re.IGNORECASE)


def parse_sheet_name(name: str, config) -> dict | None:
    """Return {facility, system, sheet_year} for a recognized sheet, else None (skip + report)."""
    if name is None:
        return None
    stripped = name.strip()
    lowered = stripped.lower()
    for facility in config.fish_care_facilities:
        if lowered.startswith(facility.lower()):
            remainder = stripped[len(facility):].strip()
            match = _SHEET_SUFFIX_RE.match(remainder)
            if not match:
                return None
            system = f"SYS {match.group(2)}" if match.group(2) else None
            return {'facility': facility, 'system': system, 'sheet_year': 2000 + int(match.group(1))}
    return None


def _normalize_header(header: list, config) -> tuple[list, list]:
    """Map each header cell to a canonical column key (or None). Unmapped non-blank headers are flagged."""
    aliases = config.fish_care_header_aliases
    col_keys, flags = [], []
    for cell in header:
        if cell is None or str(cell).strip() == '':
            col_keys.append(None)
            continue
        key = aliases.get(str(cell).strip().lower())
        if key is None:
            flags.append(f"Unmapped column '{cell}'")
        col_keys.append(key)
    return col_keys, flags


def _clean_cell(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _to_iso_date(value):
    """Normalize a date cell (openpyxl datetime, or a string) to 'YYYY-MM-DD', or None if unparseable."""
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _to_int(value):
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


def rows_from_tab(header: list, data_rows: list, meta: dict, config) -> tuple[list, list]:
    """Normalize a tab's rows into canonical row dicts. Collects validation flags (bad date,
    non-int morts, unmapped columns) without dropping the row."""
    col_keys, flags = _normalize_header(header, config)
    label = f"{meta['facility']}{(' ' + meta['system']) if meta['system'] else ''} {meta['sheet_year']}"
    out_rows = []
    for r_idx, raw in enumerate(data_rows):
        row = {}
        for c_idx, key in enumerate(col_keys):
            if key is None:
                continue
            row[key] = _clean_cell(raw[c_idx] if c_idx < len(raw) else None)
        if all(v is None for v in row.values()):
            continue  # blank row

        if row.get('obs_date') is not None:
            iso = _to_iso_date(row['obs_date'])
            if iso is None:
                flags.append(f"{label} row {r_idx + 2}: unparseable date {row['obs_date']!r}")
            row['obs_date'] = iso
        if row.get('morts') is not None:
            morts = _to_int(row['morts'])
            if morts is None:
                flags.append(f"{label} row {r_idx + 2}: non-integer morts {row['morts']!r}")
            row['morts'] = morts
        out_rows.append(row)
    return out_rows, flags


def parse_workbook(sheets: dict, config) -> dict:
    """Source-agnostic. sheets = {sheet_name: (header, data_rows)}.
    Returns {parsed: [{**meta, sheet_name, rows}], skipped: [names], flags: [str]}."""
    parsed, skipped, all_flags = [], [], []
    for name, (header, data_rows) in sheets.items():
        meta = parse_sheet_name(name, config)
        if meta is None:
            skipped.append(name)
            continue
        rows, flags = rows_from_tab(header, data_rows, meta, config)
        all_flags.extend(flags)
        parsed.append({**meta, 'sheet_name': name, 'rows': rows})
    return {'parsed': parsed, 'skipped': skipped, 'flags': all_flags}


def build_preview(parse_result: dict) -> dict:
    """A committable-summary view of a parse for the sync preview step (no DB writes)."""
    return {
        'sheets': [{
            'sheet_name': p['sheet_name'],
            'facility': p['facility'],
            'system': p['system'],
            'sheet_year': p['sheet_year'],
            'row_count': len(p['rows']),
            'sample': p['rows'][:5],
        } for p in parse_result['parsed']],
        'skipped': parse_result['skipped'],
        'flags': parse_result['flags'],
        'total_rows': sum(len(p['rows']) for p in parse_result['parsed']),
    }


def parse_xlsx_file(path: str) -> dict:
    """xlsx adapter: build the source-agnostic sheets dict from a workbook file."""
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    try:
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = [c if c is None else str(c).strip() for c in rows[0]]
            sheets[worksheet.title] = (header, [list(r) for r in rows[1:]])
    finally:
        workbook.close()
    return sheets


def import_fish_care_xlsx(dir_name: str, username: str, job_id: str):
    """Async job entry for the xlsx upload fallback. Mirrors import_master's job/complete_job contract."""
    try:
        config = get_species_config()
        path = os.path.join(dir_name, f'bulk_upload_{job_id}')
        result = parse_workbook(parse_xlsx_file(path), config)
        persist_result = persist_sheets(result['parsed'], username, 'xlsx')
        persist_result['skipped'] = result['skipped']
        persist_result['flags'] = result['flags']
        complete_job(job_id, JobState.Complete.name, persist_result)
    except Exception as e:  # noqa
        LOGGER.exception("Fish care xlsx import failed")
        complete_job(job_id, JobState.Failed.name, {"error": str(e)})
